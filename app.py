import time
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

ultima_actualizacion = datetime.now(ZoneInfo("America/Mexico_City")).strftime(
    "%d/%m/%Y %H:%M"
)

# ==========================================
# CONFIGURACIÓN GENERAL Y ESTILOS
# ==========================================
st.set_page_config(
    page_title="Quiniela Mundial 2026", page_icon="⚽", layout="wide"
)

st.markdown(
    """
    <style>
    /* Estilos Fase de Grupos */
    .stDataFrame div[data-testid="stTable"] td, 
    .stDataFrame div[data-testid="stTable"] th,
    div[data-testid="stDataFrameVisualizer"] [role="gridcell"],
    div[data-testid="stDataFrameVisualizer"] [role="columnheader"] {
        font-size: 19px !important;
    }
    .partido-hoy {
        font-size: 20px !important;
        font-weight: 500;
        margin-bottom: 8px;
    }

    /* Estilos Premium Nueva Llave Fase Final */
    .bracket-wrapper {
        background-color: #050b14;
        padding: 30px 20px;
        border-radius: 16px;
        color: #ffffff;
        font-family: 'Arial', sans-serif;
        box-shadow: 0 20px 40px rgba(0,0,0,0.5);
    }
    .bracket-header {
        text-align: center;
        margin-bottom: 30px;
    }
    .bracket-header h1 {
        color: #edd37f;
        font-size: 34px;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 2px;
        margin: 0;
    }
    .bracket-header p {
        color: #94a3b8;
        font-size: 15px;
        margin-top: 5px;
    }
    .bracket-container {
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 10px;
        overflow-x: auto;
        padding-bottom: 15px;
    }
    .bracket-column {
        display: flex;
        flex-direction: column;
        justify-content: space-around;
        height: 850px;
        min-width: 150px;
    }
    .match-box {
        display: flex;
        flex-direction: column;
        background: rgba(15, 23, 42, 0.8);
        border: 1px solid #1e293b;
        border-radius: 6px;
        padding: 2px 0;
        box-shadow: 0 4px 6px rgba(0,0,0,0.3);
    }
    .match-meta {
        font-size: 11px;
        color: #94a3b8;
        margin-bottom: 2px;
        padding-left: 4px;
        font-style: italic;
    }
    .team-row {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 6px 10px;
        font-size: 13px;
        font-weight: 600;
        letter-spacing: 0.5px;
        color: #cbd5e1;
        border-bottom: 1px solid rgba(51, 65, 85, 0.4);
    }
    .team-row:last-child {
        border-bottom: none;
    }
    .center-trophy {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        min-width: 200px;
        text-align: center;
    }
    .trophy-title {
        color: #edd37f;
        font-size: 18px;
        font-weight: bold;
        letter-spacing: 1px;
        margin-bottom: 12px;
    }
    .champion-display {
        background: linear-gradient(135deg, #dfb743 0%, #b8860b 100%);
        color: #050b14;
        padding: 15px 22px;
        border-radius: 8px;
        font-weight: 800;
        font-size: 17px;
        box-shadow: 0 0 25px rgba(223, 183, 67, 0.4);
        border: 1px solid #fef08a;
        margin-top: 15px;
        text-transform: uppercase;
    }
    </style>
    """,
    unsafe_allow_html=True
)

inicio = time.time()

st.title("⚽ Quiniela Mundial 2026")
st.info("⚽ La información se actualiza desde Google Drive. La carga inicial puede tardar algunos segundos.")
st.caption(f"Página actualizada: {ultima_actualizacion} (hora CDMX)")

# ==========================================
# 🔐 FILTRO DE PESTAÑAS (MODO OCULTO PARA DESARROLLO)
# ==========================================
menu_opciones = [
    "🏆 Ranking", 
    "👤 Participantes", 
    "⚽ Partidos", 
    "🔥 Comparativa VS", 
    "🗓️ Calendario"
]

query_params = st.query_params
if query_params.get("dev") == "true":
    menu_opciones.append("🗓️ Llave Fase Final (Prueba)")

seleccion_menu = st.sidebar.radio("Menú", menu_opciones)

# Normalizamos la selección eliminando emojis para evitar fallos de renderizado
if "Ranking" in seleccion_menu: pagina = "Ranking"
elif "Participantes" in seleccion_menu: pagina = "Participantes"
elif "Partidos" in seleccion_menu: pagina = "Partidos"
elif "Comparativa" in seleccion_menu: pagina = "Comparativa VS"
elif "Calendario" in seleccion_menu: pagina = "Calendario"
elif "Llave" in seleccion_menu: pagina = "Llave Fase Final"
else: pagina = "Ranking"

# ==========================================
# CONEXIÓN Y PROCESAMIENTO DE GOOGLE DRIVE
# ==========================================
FILE_ID = "1svfBlcw4oOEltibwpv1c8I4h6sHmeq7z"
URL_DRIVE = f"https://docs.google.com/uc?export=download&id={FILE_ID}"

@st.cache_data(ttl=60)
def cargar_excel():
    try:
        respuesta = requests.get(URL_DRIVE, timeout=10)
        if respuesta.status_code == 200:
            return respuesta.content
    except Exception as e:
        st.error(f"Error al descargar Excel de Drive: {e}")
    return None

def determinar_resultado_celdas(c, d, e):
    c_str = str(c).strip().lower() if c is not None else ""
    d_str = str(d).strip().lower() if d is not None else ""
    e_str = str(e).strip().lower() if e is not None else ""
    if c_str == "x": return "Local"
    elif d_str == "x": return "Empate"
    elif e_str == "x": return "Visitante"
    return None

@st.cache_data(ttl=60)  
def procesar_todo_el_excel(contenido_excel):
    wb_local = load_workbook(BytesIO(contenido_excel), data_only=True, read_only=True)
    if "RESULTADOS" not in wb_local.sheetnames:
        return None, None

    ws_resultados = wb_local["RESULTADOS"]
    resultados_oficiales = {}
    
    for fila_idx, row in enumerate(ws_resultados.iter_rows(min_row=6, max_row=500, min_col=2, max_col=6, values_only=True), start=6):
        if len(row) < 5: continue
        local, c, d, e, visitante = row[0], row[1], row[2], row[3], row[4]
        if local is None or visitante is None: continue
        resultados_oficiales[fila_idx] = determinar_resultado_celdas(c, d, e)

    participantes_local = {}
    calendario_local = []

    for hoja in wb_local.sheetnames:
        if hoja.upper() in ["RESULTADOS", "CALENDARIO", "MURO"]: continue

        ws = wb_local[hoja]
        nombre = hoja
        desempate_local = "-"
        desempate_visitante = "-"
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=15, min_col=3, max_col=12, values_only=True), start=2):
            if r_idx == 2 and len(row) > 0: nombre = row[0] or hoja
            if r_idx == 15 and len(row) >= 10:
                desempate_local = row[7]       
                desempate_visitante = row[9]   

        pronosticos = []
        for fila_idx, row in enumerate(ws.iter_rows(min_row=6, max_row=500, min_col=2, max_col=6, values_only=True), start=6):
            if len(row) < 5: continue
            local, c, d, e, visitante = row[0], row[1], row[2], row[3], row[4]
            if local is None or visitante is None: continue

            resultado_oficial = resultados_oficiales.get(fila_idx)
            pronostico_jugador = determinar_resultado_celdas(c, d, e)
            es_acierto = (resultado_oficial is not None and pronostico_jugador == resultado_oficial)

            estatus_visual = "⌛ Pautado"
            if resultado_oficial is not None:
                estatus_visual = "✅ ¡Acertó!" if es_acierto else "❌ Falló"

            pronosticos.append({
                "fila": fila_idx,
                "Partido": f"{str(local).strip()} vs {str(visitante).strip()}",
                "Pronóstico": pronostico_jugador,
                "Resultado Oficial": resultado_oficial,
                "Estatus": estatus_visual,
                "Acierto": es_acierto,
            })

        participantes_local[nombre] = {
            "pronosticos": pronosticos,
            "desempate_local": desempate_local,
            "desempate_visitante": desempate_visitante,
        }

    if "CALENDARIO" in wb_local.sheetnames:
        ws_cal = wb_local["CALENDARIO"]
        for row in ws_cal.iter_rows(min_row=2, max_row=500, min_col=1, max_col=4, values_only=True):
            if len(row) < 4 or row[0] is None: continue
            calendario_local.append({
                "partido": row[0], "fecha": row[1], "hora": row[2], "resultado": row[3]
            })

    return participantes_local, calendario_local

# ==========================================
# EJECUCIÓN LÓGICA DE CONTROL
# ==========================================
contenido_excel = cargar_excel()
if contenido_excel is None:
    st.error("No se pudo descargar el archivo desde Google Drive.")
    st.stop()

participantes, calendario_datos = procesar_todo_el_excel(contenido_excel)
if participantes is None:
    st.error("No existe la hoja RESULTADOS en el archivo.")
    st.stop()

def normalizar_texto_partido(texto):
    if not texto: 
        return ""
    return " ".join(str(texto).split()).lower()

orden_calendario = {}
if calendario_datos:
    for idx, c_partido in enumerate(calendario_datos):
        orden_calendario[normalizar_texto_partido(c_partido["partido"])] = idx

puntos = {nombre: sum(1 for p in datos["pronosticos"] if p["Acierto"]) for nombre, datos in participantes.items()}
st.write(f"Tiempo de carga: {round(time.time() - inicio, 2)} segundos")

# ==========================================
# RENDERIZADO DE LAS PÁGINAS ORIGINALES
# ==========================================
if pagina == "Ranking":
    ranking_datos = []
    for nombre in participantes:
        dl = participantes[nombre]['desempate_local']
        dv = participantes[nombre]['desempate_visitante']
        try:
            val_dl = f"{int(float(dl))}" if dl not in [None, "", "-"] else "-"
            val_dv = f"{int(float(dv))}" if dv not in [None, "", "-"] else "-"
            desempate_txt = f"{val_dl}-{val_dv}" if val_dl != "-" and val_dv != "-" else "-"
        except:
            desempate_txt = f"{dl}-{dv}" if dl or dv else "-"

        ranking_datos.append({
            "Participante": nombre, "Puntos": puntos[nombre], "Desempate (Chequia vs México)": desempate_txt
        })

    ranking = pd.DataFrame(ranking_datos)
    ranking = ranking.sort_values(by="Puntos", ascending=False).reset_index(drop=True)

    total_partidos, partidos_jugados, partidos_hoy = 0, 0, []
    hoy = datetime.now(ZoneInfo("America/Mexico_City")).date()

    if calendario_datos:
        for c_partido in calendario_datos:
            partido, fecha, hora, resultado = c_partido["partido"], c_partido["fecha"], c_partido["hora"], c_partido["resultado"]
            total_partidos += 1
            if resultado not in [None, ""]: partidos_jugados += 1
            try:
                if hasattr(fecha, "date") and fecha.date() == hoy:
                    if resultado not in [None, ""]:
                        equipos = partido.split(" vs ")
                        texto = f"⚽ {equipos[0]} {resultado} {equipos[1]}" if len(equipos) == 2 else f"⚽ {partido} ({resultado})"
                    else:
                        texto = f"🕒 {hora.strftime('%H:%M')} - {partido}" if hasattr(hora, "strftime") else f"{hora} - {partido}"
                    partidos_hoy.append(texto)
            except: pass

    porcentaje = round(partidos_jugados * 100 / total_partidos, 1) if total_partidos > 0 else 0
    puntaje_maximo = ranking["Puntos"].max() if not ranking.empty else -1
    puntaje_minimo = ranking["Puntos"].min() if not ranking.empty else -1
    
    if puntaje_maximo != -1:
        filtro_lideres = ranking[ranking["Puntos"] == puntaje_maximo]["Participante"].tolist()
        primeros_nombres_lideres = [str(n).strip().split()[0] for n in filtro_lideres]
        texto_lideres = ", ".join(primeros_nombres_lideres[:-1]) + " y " + primeros_nombres_lideres[-1] if len(primeros_nombres_lideres) > 1 else primeros_nombres_lideres[0]
        etiqueta_lider = "🔥 Líderes Actuales" if len(primeros_nombres_lideres) > 1 else "🔥 Líder Actual"
    else:
        texto_lideres, etiqueta_lider = "-", "🔥 Líder Actual"

    col1, col2, col3 = st.columns(3)
    with col1: st.metric(label="⚽ Partidos Jugados", value=f"{partidos_jugados} / {total_partidos}")
    with col2: st.metric(label="📈 Avance del Torneo", value=f"{porcentaje}%")
    with col3: st.metric(label=etiqueta_lider, value=texto_lideres)

    st.divider()
    st.subheader("📅 Partidos para hoy")
    if len(partidos_hoy) == 0: st.info("No hay partidos programados para hoy.")
    else:
        for p in partidos_hoy: st.markdown(f'<p class="partido-hoy">{p}</p>', unsafe_allow_html=True)

    st.divider()
    st.subheader("Tabla General")
    if puntaje_maximo != -1:
        def agregar_emoji(r):
            if str(r["Participante"]).strip() == "Victor Vazquez": return f"🐌 {r['Participante']}"
            if r["Puntos"] == puntaje_maximo: return f"👑 {r['Participante']}"
            elif r["Puntos"] == puntaje_minimo and puntaje_minimo != puntaje_maximo: return f"🐌 {r['Participante']}"
            return r["Participante"]
        ranking["Participante"] = ranking.apply(agregar_emoji, axis=1)

    def resaltar_estilo_premium(row):
        if puntaje_maximo != -1 and row["Puntos"] == puntaje_maximo: return ['background-color: #fffbeb; color: #b45309; font-weight: bold;'] * len(row)
        elif puntaje_minimo != -1 and row["Puntos"] == puntaje_minimo and puntaje_minimo != puntaje_maximo: return ['background-color: #fdf2f8; color: #9d174d; font-style: italic;'] * len(row)
        return [''] * len(row)

    st.dataframe(ranking.style.apply(resaltar_estilo_premium, axis=1), use_container_width=True, hide_index=True)

elif pagina == "Participantes":
    jugador = st.selectbox("Selecciona participante", list(participantes.keys()))
    st.subheader(f"Pronósticos de {jugador}")
    df = pd.DataFrame(participantes[jugador]["pronosticos"])
    df["_orden"] = df["Partido"].map(lambda x: orden_calendario.get(normalizar_texto_partido(x), 999))
    df = df.sort_values(by="_orden", ascending=True).reset_index(drop=True)[["Partido", "Pronóstico", "Resultado Oficial", "Estatus"]]

    def color_estatus(val):
        if "✅" in str(val): return 'background-color: #d4edda; color: #155724; font-weight: bold;'
        elif "❌" in str(val): return 'background-color: #f8d7da; color: #721c24;'
        return ''
    st.dataframe(df.style.map(color_estatus, subset=["Estatus"]), use_container_width=True, hide_index=True)

elif pagina == "Partidos":
    st.subheader("⚽ Análisis por Partido")
    if calendario_datos:
        lista_partidos = [c_partido["partido"] for c_partido in calendario_datos]
    else:
        primer_jugador = list(participantes.keys())[0]
        lista_partidos_cruda = [p["Partido"] for p in participantes[primer_jugador]["pronosticos"]]
        lista_partidos = sorted(lista_partidos_cruda, key=lambda x: orden_calendario.get(normalizar_texto_partido(x), 999))
    
    partido_seleccionado = st.selectbox("Selecciona un partido para ver las predicciones:", lista_partidos)
    datos_partido = []

    for nombre, datos in participantes.items():
        for p in datos["pronosticos"]:
            if normalizar_texto_partido(p["Partido"]) == normalizar_texto_partido(partido_seleccionado):
                datos_partido.append({
                    "Participante": nombre,
                    "Pronóstico": p["Pronóstico"],
                    "Resultado Oficial": p["Resultado Oficial"] if p["Resultado Oficial"] else "⌛ Pautado",
                    "¿Acertó?": "✅ SÍ" if p["Acierto"] else "❌ NO",
                })
    st.dataframe(pd.DataFrame(datos_partido), use_container_width=True, hide_index=True)

elif pagina == "Comparativa VS":
    st.subheader("🥊 Cara a Cara entre Participantes")
    seleccionados = st.multiselect("Selecciona de 2 a 3 participantes para el Versus:", options=list(participantes.keys()), max_selections=3)
    if len(seleccionados) < 2:
        st.info("💡 Selecciona al menos 2 participantes en el cuadro de arriba para generar el frente a frente.")
    else:
        primer_p = seleccionados[0]
        datos_vs = []
        partidos_lista = [p["Partido"] for p in participantes[primer_p]["pronosticos"]]
        for i, p_nombre in enumerate(partidos_lista):
            fila_vs = {"Partido": p_nombre}
            resultado_real = participantes[primer_p]["pronosticos"][i]["Resultado Oficial"]
            fila_vs["Resultado Real"] = resultado_real if resultado_real else "⌛ Pautado"
            for jug in seleccionados: fila_vs[f"Pred. {jug}"] = participantes[jug]["pronosticos"][i]["Pronóstico"]
            fila_vs["_orden"] = orden_calendario.get(normalizar_texto_partido(p_nombre), 999)
            datos_vs.append(fila_vs)
            
        df_vs = pd.DataFrame(datos_vs).sort_values(by="_orden", ascending=True).drop(columns=["_orden"]).reset_index(drop=True)
        st.dataframe(df_vs.style.map(lambda v: 'font-weight: bold;' if v in ["Local", "Empate", "Visitante"] else ('color: #888888; font-style: italic;' if "⌛" in str(v) else '')), use_container_width=True, hide_index=True)

elif pagina == "Calendario":
    if not calendario_datos: st.warning("No existe o está vacía la hoja CALENDARIO")
    else:
        calendario_tabla = []
        for c_partido in calendario_datos:
            partido, fecha, hora, rf = c_partido["partido"], c_partido["fecha"], c_partido["hora"], c_partido["resultado"]
            try: fecha = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)
            except: pass
            try: hora = hora.strftime("%H:%M") if hasattr(hora, "strftime") else str(hora)
            except: pass
            calendario_tabla.append({"Partido": partido, "Fecha": fecha, "Hora (CDMX)": hora, "Resultado Final": rf if rf is not None else " "})
        st.subheader("Calendario de partidos")
        st.dataframe(pd.DataFrame(calendario_tabla), use_container_width=True, hide_index=True)

# ==========================================
# 🆕 PESTAÑA OCULTA: LLAVE FASE FINAL (ESTRUCTURA PURA)
# ==========================================
elif pagina == "Llave Fase Final":
    def render_match_html(id_partido, meta_text=""):
        return f"""<div><div class="match-meta">{meta_text}</div><div class="match-box"><div class="team-row"><span>🏳️ Por clasificar</span></div><div class="team-row"><span>🏳️ Por clasificar</span></div></div></div>"""

    st.markdown("""<div class="bracket-header"><h1>FLUJO OFICIAL FASE FINAL</h1><p>Estructura de llaves y sedes del torneo desde Dieciseisavos hasta la Gran Final</p></div>""", unsafe_allow_html=True)

    html_llave = f"""
<div class="bracket-wrapper">
    <div class="bracket-container">
        <!-- === BLOQUE IZQUIERDO === -->
        <div class="bracket-column">
            {render_match_html("D16_1", "28/06 Los Ángeles")}
            {render_match_html("D16_2", "29/06 Boston")}
            {render_match_html("D16_3", "29/06 Monterrey")}
            {render_match_html("D16_4", "29/06 Houston")}
            {render_match_html("D16_5", "30/06 NY/NJ")}
            {render_match_html("D16_6", "30/06 Dallas")}
            {render_match_html("D16_7", "30/06 CDMX")}
            {render_match_html("D16_8", "01/07 Atlanta")}
        </div>
        <div class="bracket-column">
            {render_match_html("OCT_1", "04/07 Filadelfia")}
            {render_match_html("OCT_2", "04/07 Houston")}
            {render_match_html("OCT_3", "06/07 Dallas")}
            {render_match_html("OCT_4", "06/07 CDMX")}
        </div>
        <div class="bracket-column">
            {render_match_html("CRT_1", "09/07 Boston")}
            {render_match_html("CRT_2", "10/07 Los Ángeles")}
        </div>
        <div class="bracket-column">
            {render_match_html("SEM_1", "14/07 Dallas")}
        </div>
        <!-- === CENTRO COPA === -->
        <div class="center-trophy">
            <div class="trophy-title">19/07 Nueva York</div>
            <div>
                <div class="match-meta">GRAN FINAL</div>
                <div class="match-box">
                    <div class="team-row"><span>🏳️ Finalista A</span></div>
                    <div class="team-row"><span>🏳️ Finalista B</span></div>
                </div>
            </div>
            <div class="champion-display">
                <div style="font-size: 10px; font-weight: bold; opacity: 0.9; letter-spacing: 1px;">CAMPEÓN MUNDIAL</div>
                <div>🏆 ⌛</div>
            </div>
        </div>
        <!-- === BLOQUE DERECHO === -->
        <div class="bracket-column">
            {render_match_html("SEM_2", "15/07 Atlanta")}
        </div>
        <div class="bracket-column">
            {render_match_html("CRT_3", "11/07 Miami")}
            {render_match_html("CRT_4", "11/07 Kansas City")}
        </div>
        <div class="bracket-column">
            {render_match_html("OCT_5", "05/07 CDMX")}
            {render_match_html("OCT_6", "05/07 Nueva York")}
            {render_match_html("OCT_7", "07/07 Atlanta")}
            {render_match_html("OCT_8", "07/07 Vancouver")}
        </div>
        <div class="bracket-column">
            {render_match_html("D16_9", "01/07 S. Francisco")}
            {render_match_html("D16_10", "01/07 Seattle")}
            {render_match_html("D16_11", "02/07 Toronto")}
            {render_match_html("D16_12", "02/07 Los Ángeles")}
            {render_match_html("D16_13", "02/07 Vancouver")}
            {render_match_html("D16_14", "03/07 Miami")}
            {render_match_html("D16_15", "03/07 Kansas City")}
            {render_match_html("D16_16", "03/07 Dallas")}
        </div>
    </div>
</div>
"""
    st.html(html_llave)
